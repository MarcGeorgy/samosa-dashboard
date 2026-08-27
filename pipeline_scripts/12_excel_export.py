"""
Builds a styled, multi-sheet Excel workbook from the pipeline's current
output - the "export to Excel" button in 11_debrief_app.py, and also
runnable standalone for a one-off export without the dashboard.

Sheets:
  1. Overview            headline metrics + a quality-band bar chart
  2. Flagged Records      every flagged submission, rows colored by band
  3. Enumerator Summary   per-enumerator QA rollup, colored by avg score
  4. Daily Trend          per-day rollup + a trend line chart
  5. Comments             enumerator comments with LLM tags/severity

Usage: build_workbook(flagged, enum_summary, daily, comments, summary_json) -> bytes
"""
import importlib.util
import io
import json
import sys
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent
OUT_DIR = Path(tempfile.gettempdir()) / "mvsy_monitoring" / "output"

# ------------------------------------------------------------------ styling
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
SUBTITLE_FONT = Font(italic=True, size=10, color="666666")
THIN_BORDER = Border(*[Side(style="thin", color="D9D9D9")] * 4)

BAND_FILLS = {
    "A - clean": PatternFill("solid", fgColor="C6EFCE"),
    "B - minor issues": PatternFill("solid", fgColor="FFEB9C"),
    "C - needs review": PatternFill("solid", fgColor="FFD08A"),
    "D - flag for supervisor callback": PatternFill("solid", fgColor="FFC7CE"),
}
BAND_FONTS = {
    "A - clean": Font(color="006100"),
    "B - minor issues": Font(color="9C6500"),
    "C - needs review": Font(color="9C5700"),
    "D - flag for supervisor callback": Font(color="9C0006"),
}


def _write_title(ws, text, subtitle=None):
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUBTITLE_FONT


def _write_table(ws, df: pd.DataFrame, start_row: int, band_col: str | None = None,
                  wrap_cols: list | None = None) -> int:
    """Writes df as a styled table starting at start_row (1-indexed). Returns the row after the table."""
    header_row = start_row
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=str(col_name))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    columns = list(df.columns)
    band_col_idx = columns.index(band_col) + 1 if band_col and band_col in columns else None
    wrap_col_idxs = {columns.index(c) + 1 for c in (wrap_cols or []) if c in columns}
    for r_offset, (_, row) in enumerate(df.iterrows(), start=1):
        r = header_row + r_offset
        band_val = row.get(band_col) if band_col_idx else None
        for c_idx, val in enumerate(row, start=1):
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c_idx in wrap_col_idxs))
            if band_val in BAND_FILLS:
                cell.fill = BAND_FILLS[band_val]
                cell.font = BAND_FONTS[band_val]

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(df.columns))}{header_row + len(df)}"

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str)][:200])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)

    return header_row + len(df) + 2


def _metrics_block(ws, start_row: int, pairs: list[tuple]) -> int:
    r = start_row
    for label, value in pairs:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        r += 1
    return r + 1


# ------------------------------------------------------------------ sheets
def _overview_sheet(wb, summary_json: dict):
    ws = wb.active
    ws.title = "Overview"
    _write_title(ws, "MSY Listing Survey - Data Quality Overview",
                 f"Generated {summary_json.get('generated_at', '')}")

    r = _metrics_block(ws, 4, [
        ("Total submissions", summary_json.get("n_submissions")),
        ("Completed interviews", summary_json.get("n_completed")),
        ("Enumerators", summary_json.get("n_enumerators")),
        ("Villages", summary_json.get("n_villages")),
        ("Average quality score", summary_json.get("avg_quality_score")),
        ("% of submissions flagged", f"{summary_json.get('pct_flagged_any')}%"),
        ("Exact duplicates", summary_json.get("n_exact_duplicates")),
        ("Fuzzy duplicates", summary_json.get("n_fuzzy_duplicates")),
        ("Statistical outlier rows", summary_json.get("n_outlier_rows")),
        ("Enumerator-fatigue flags", summary_json.get("n_enumerator_fatigue_flags")),
        ("Respondent-fatigue flags", summary_json.get("n_respondent_fatigue_flags")),
        ("Off-hours submissions (before 7am / after 7pm)", summary_json.get("n_off_hours_flags")),
        ("Possible fake/fabricated entries (implausibly fast)", summary_json.get("n_possible_fake_entries")),
        ("Submissions on an outdated form version", summary_json.get("n_stale_form_version")),
        ("Enumerator comments received", summary_json.get("n_comments")),
    ])

    band_counts = summary_json.get("quality_band_counts", {})
    band_order = ["A - clean", "B - minor issues", "C - needs review", "D - flag for supervisor callback"]
    ws.cell(row=r, column=1, value="Quality band").font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    ws.cell(row=r, column=2, value="Count").font = HEADER_FONT
    ws.cell(row=r, column=2).fill = HEADER_FILL
    chart_start = r
    for i, band in enumerate(band_order, start=1):
        ws.cell(row=r + i, column=1, value=band)
        ws.cell(row=r + i, column=2, value=band_counts.get(band, 0))
        fill = BAND_FILLS.get(band)
        if fill:
            ws.cell(row=r + i, column=1).fill = fill

    chart = BarChart()
    chart.title = "Submissions by quality band"
    chart.y_axis.title = "Submissions"
    data = Reference(ws, min_col=2, min_row=chart_start, max_row=chart_start + len(band_order))
    cats = Reference(ws, min_col=1, min_row=chart_start + 1, max_row=chart_start + len(band_order))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width, chart.height = 16, 9
    ws.add_chart(chart, "D4")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16


def _flagged_sheet(wb, flagged: pd.DataFrame):
    ws = wb.create_sheet("Flagged Records")
    _write_title(ws, "Flagged Submissions", "Every submission with 1+ quality flag, colored by quality band")
    df = flagged[flagged["n_flags"].astype(float) > 0].copy()
    display_cols = ["enumerator_id", "district", "village", "hh_id", "hh_outcome", "duration_min",
                     "quality_band", "data_quality_score", "n_flags"] + \
                    [c for c in df.columns if c.startswith("flag_")]
    display_cols = [c for c in display_cols if c in df.columns]
    _write_table(ws, df[display_cols].sort_values("data_quality_score"), start_row=4, band_col="quality_band")


def _enum_summary_sheet(wb, enum_summary: pd.DataFrame):
    ws = wb.create_sheet("Enumerator Summary")
    _write_title(ws, "Enumerator Performance Summary")
    df = enum_summary.copy()

    def score_band(s):
        if s >= 90:
            return "A - clean"
        if s >= 75:
            return "B - minor issues"
        if s >= 55:
            return "C - needs review"
        return "D - flag for supervisor callback"
    df["_band"] = df["avg_quality_score"].apply(score_band)
    _write_table(ws, df, start_row=4, band_col="_band")


def _daily_sheet(wb, daily: pd.DataFrame):
    ws = wb.create_sheet("Daily Trend")
    _write_title(ws, "Daily Data Quality Trend")
    end_row = _write_table(ws, daily, start_row=4)

    n = len(daily)
    score_col = list(daily.columns).index("avg_quality_score") + 1
    chart = LineChart()
    chart.title = "Average quality score by day"
    chart.y_axis.title = "Avg quality score"
    data = Reference(ws, min_col=score_col, min_row=4, max_row=4 + n)
    cats = Reference(ws, min_col=1, min_row=5, max_row=4 + n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width, chart.height = 18, 9
    ws.add_chart(chart, f"A{end_row}")


def _comments_sheet(wb, comments: pd.DataFrame):
    ws = wb.create_sheet("Comments")
    _write_title(ws, "Enumerator Field Comments")
    cols = [c for c in ["SubmissionDate", "enumerator_id", "village", "hh_id", "hh_outcome",
                         "comment_tag", "llm_tags", "llm_severity", "llm_recommended_action",
                         "enumerator_comments", "data_quality_score"] if c in comments.columns]
    _write_table(ws, comments[cols], start_row=4, wrap_cols=["enumerator_comments", "llm_recommended_action"])
    if "enumerator_comments" in cols:
        ws.column_dimensions[get_column_letter(cols.index("enumerator_comments") + 1)].width = 60


# -------------------------------------------------------------------- API
def build_workbook(flagged: pd.DataFrame, enum_summary: pd.DataFrame, daily: pd.DataFrame,
                    comments: pd.DataFrame, summary_json: dict) -> bytes:
    # Callers may pass string-dtype frames (e.g. the dashboard loads `flagged`
    # with dtype=str so id-like columns stay exact) - coerce the numeric
    # columns here so sorting and the Excel charts see real numbers, not text.
    flagged = flagged.copy()
    for col in ["data_quality_score", "n_flags", "duration_min"]:
        if col in flagged.columns:
            flagged[col] = pd.to_numeric(flagged[col], errors="coerce")

    enum_summary = enum_summary.copy()
    for col in enum_summary.columns:
        if col != "enumerator_id":
            enum_summary[col] = pd.to_numeric(enum_summary[col], errors="coerce")

    daily = daily.copy()
    for col in daily.columns:
        if col != "work_date":
            daily[col] = pd.to_numeric(daily[col], errors="coerce")

    comments = comments.copy()
    if "data_quality_score" in comments.columns:
        comments["data_quality_score"] = pd.to_numeric(comments["data_quality_score"], errors="coerce")

    wb = Workbook()
    _overview_sheet(wb, summary_json)
    _flagged_sheet(wb, flagged)
    _enum_summary_sheet(wb, enum_summary)
    _daily_sheet(wb, daily)
    _comments_sheet(wb, comments)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    flagged = pd.read_csv(OUT_DIR / "msy_listing_flagged.csv", dtype=str, keep_default_na=False)
    for col in ["data_quality_score", "n_flags", "duration_min"]:
        flagged[col] = pd.to_numeric(flagged[col], errors="coerce")
    enum_summary = pd.read_csv(OUT_DIR / "enumerator_summary.csv")
    daily = pd.read_csv(OUT_DIR / "daily_summary.csv")
    comments = pd.read_csv(OUT_DIR / "comments_feed.csv")
    summary_json = json.loads((OUT_DIR / "monitoring_summary.json").read_text())

    xlsx_bytes = build_workbook(flagged, enum_summary, daily, comments, summary_json)
    out_path = OUT_DIR / "msy_dashboard_export.xlsx"
    out_path.write_bytes(xlsx_bytes)
    print(f"wrote {out_path} ({len(xlsx_bytes):,} bytes)")
